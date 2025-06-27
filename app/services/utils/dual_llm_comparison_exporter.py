import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List, Dict, Any, Optional
import json
import os
from datetime import datetime

from app.models.screening_models import Article


class DualLLMComparisonExporter:
    """
    Exports dual-LLM screening results to color-coded Excel spreadsheets for comparison analysis.
    """
    
    def __init__(self):
        self.openai_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
        self.anthropic_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light Pink
        self.agreement_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light Green
        self.disagreement_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light Yellow
        
        self.header_font = Font(bold=True)
        self.center_alignment = Alignment(horizontal="center", vertical="center")
    
    def extract_first_author_year(self, article: Article) -> str:
        """Extract first author and year for study identification."""
        authors = article.authors or ""
        first_author = authors.split(',')[0].strip() if authors else "Unknown"
        year = str(article.year) if article.year else "Unknown"
        return f"{first_author}, {year}"
    
    def extract_pico_elements(self, result: Dict[str, Any]) -> Dict[str, str]:
        """Extract PICO-TT elements with exact quotes."""
        pico = result.get('picott_extraction', {})
        return {
            'Population': ', '.join(pico.get('population', [])) if pico.get('population') else 'Not found',
            'Intervention': ', '.join(pico.get('intervention', [])) if pico.get('intervention') else 'Not found',
            'Comparison': ', '.join(pico.get('comparison', [])) if pico.get('comparison') else 'Not found',
            'Outcomes': ', '.join(pico.get('outcomes', [])) if pico.get('outcomes') else 'Not found',
            'Time': pico.get('time_frame', 'Not found'),
            'Study_Type': pico.get('study_type', 'Not found')
        }
    
    def parse_decision_reasoning(self, decision_reasoning: str) -> Dict[str, Any]:
        """Parse the JSON decision_reasoning field to extract dual-LLM results."""
        try:
            data = json.loads(decision_reasoning)
            return data
        except (json.JSONDecodeError, TypeError):
            return {}
    
    def extract_llm_results(self, decision_data: Dict[str, Any]) -> tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
        """Extract OpenAI and Anthropic results from decision data."""
        openai_result = decision_data.get('openai_result')
        anthropic_result = decision_data.get('anthropic_result')
        
        return openai_result, anthropic_result
    
    def determine_agreement_status(self, openai_result: Optional[Dict[str, Any]], 
                                 anthropic_result: Optional[Dict[str, Any]]) -> str:
        """Determine if the two LLMs agree on the screening decision."""
        if not openai_result or not anthropic_result:
            return "Incomplete"
        
        openai_decision = openai_result.get('screening_decision', {}).get('final_decision')
        anthropic_decision = anthropic_result.get('screening_decision', {}).get('final_decision')
        
        if openai_decision == anthropic_decision:
            return "Agreement"
        else:
            return "Disagreement"
    
    def calculate_agreement_metrics(self, openai_result: Optional[Dict[str, Any]], 
                                  anthropic_result: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate advanced agreement metrics based on RealDualLLMEvaluator logic."""
        if not openai_result or not anthropic_result:
            return {
                'decision_agreement': False,
                'confidence_delta': 100,
                'conflict_detected': True,
                'low_confidence': True,
                'agreement_score': 0.0,
                'openai_confidence': 0,
                'anthropic_confidence': 0,
                'notes': 'Incomplete data'
            }
        
        openai_decision = openai_result.get('screening_decision', {}).get('final_decision')
        anthropic_decision = anthropic_result.get('screening_decision', {}).get('final_decision')
        openai_conf = openai_result.get('screening_decision', {}).get('confidence_score', 0)
        anthropic_conf = anthropic_result.get('screening_decision', {}).get('confidence_score', 0)
        
        # Decision agreement
        decision_agreement = openai_decision == anthropic_decision
        
        # Confidence delta
        confidence_delta = abs(openai_conf - anthropic_conf)
        
        conflict_detected = not decision_agreement or confidence_delta > 20
        
        # Low confidence detection
        low_confidence = min(openai_conf, anthropic_conf) < 70
        
        # Overall agreement score
        if decision_agreement:
            agreement_score = 1.0 - (confidence_delta / 100)
        else:
            agreement_score = 0.0
        
        return {
            'decision_agreement': decision_agreement,
            'confidence_delta': confidence_delta,
            'conflict_detected': conflict_detected,
            'low_confidence': low_confidence,
            'agreement_score': agreement_score,
            'openai_confidence': openai_conf,
            'anthropic_confidence': anthropic_conf,
            'requires_human_review': conflict_detected or low_confidence
        }
    
    def resolve_final_decision(self, openai_result: Optional[Dict[str, Any]], 
                             anthropic_result: Optional[Dict[str, Any]], 
                             agreement_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Resolve final decision based on LLM evaluations using RealDualLLMEvaluator logic."""
        if not openai_result or not anthropic_result:
            return {
                'decision': 'INCOMPLETE',
                'confidence': 0,
                'reasoning': 'Missing evaluation data',
                'resolution_method': 'error'
            }
        
        if agreement_metrics['conflict_detected']:
            return {
                'decision': 'CONFLICT',
                'confidence': min(agreement_metrics['openai_confidence'], agreement_metrics['anthropic_confidence']),
                'reasoning': 'LLM disagreement detected - requires human review',
                'resolution_method': 'human_review_required'
            }
        
        # If agreement, use higher confidence evaluation
        if agreement_metrics['openai_confidence'] >= agreement_metrics['anthropic_confidence']:
            primary_eval = openai_result
            secondary_eval = anthropic_result
            primary_provider = 'OpenAI'
        else:
            primary_eval = anthropic_result
            secondary_eval = openai_result
            primary_provider = 'Anthropic'
        
        primary_decision = primary_eval.get('screening_decision', {}).get('final_decision', 'UNCERTAIN')
        primary_reasoning = primary_eval.get('screening_decision', {}).get('detailed_reasoning', 'N/A')
        
        return {
            'decision': primary_decision,
            'confidence': (agreement_metrics['openai_confidence'] + agreement_metrics['anthropic_confidence']) / 2,
            'reasoning': f"Agreement reached via {primary_provider}. {primary_reasoning[:200]}...",
            'resolution_method': 'llm_consensus',
            'primary_provider': primary_provider
        }
    
    def calculate_quality_metrics(self, articles: List[Article]) -> Dict[str, Any]:
        """Calculate quality metrics for the entire dataset."""
        total_evaluations = 0
        conflicts = 0
        human_reviews = 0
        agreement_scores = []
        confidence_scores = []
        
        decision_counts = {'INCLUDE': 0, 'EXCLUDE': 0, 'CONFLICT': 0, 'INCOMPLETE': 0}
        
        for article in articles:
            if not article.decision_reasoning:
                continue
                
            decision_data = self.parse_decision_reasoning(article.decision_reasoning)
            if not decision_data:
                continue
                
            openai_result, anthropic_result = self.extract_llm_results(decision_data)
            if not openai_result or not anthropic_result:
                continue
                
            total_evaluations += 1
            
            agreement_metrics = self.calculate_agreement_metrics(openai_result, anthropic_result)
            final_decision = self.resolve_final_decision(openai_result, anthropic_result, agreement_metrics)
            
            if agreement_metrics['conflict_detected']:
                conflicts += 1
            if agreement_metrics['requires_human_review']:
                human_reviews += 1
                
            agreement_scores.append(agreement_metrics['agreement_score'])
            confidence_scores.extend([agreement_metrics['openai_confidence'], agreement_metrics['anthropic_confidence']])
            
            decision = final_decision['decision']
            if decision in decision_counts:
                decision_counts[decision] += 1
        
        avg_agreement = sum(agreement_scores) / len(agreement_scores) if agreement_scores else 0
        avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0
        
        return {
            'total_evaluations': total_evaluations,
            'decision_distribution': decision_counts,
            'quality_metrics': {
                'conflict_rate': (conflicts / total_evaluations * 100) if total_evaluations > 0 else 0,
                'human_review_rate': (human_reviews / total_evaluations * 100) if total_evaluations > 0 else 0,
                'average_confidence': avg_confidence,
                'average_agreement': avg_agreement * 100
            }
        }
    
    def generate_comparison_spreadsheet(self, articles: List[Article], project_name: str) -> str:
        """Generate Excel spreadsheet with dual-LLM comparison."""
        
        comparison_data = []
        
        for article in articles:
            if not article.decision_reasoning:
                continue
            
            decision_data = self.parse_decision_reasoning(article.decision_reasoning)
            if not decision_data:
                continue
            
            openai_result, anthropic_result = self.extract_llm_results(decision_data)
            
            if not openai_result or not anthropic_result:
                continue
            
            study_id = self.extract_first_author_year(article)
            agreement_status = self.determine_agreement_status(openai_result, anthropic_result)
            
            agreement_metrics = self.calculate_agreement_metrics(openai_result, anthropic_result)
            final_decision = self.resolve_final_decision(openai_result, anthropic_result, agreement_metrics)
            
            openai_pico = self.extract_pico_elements(openai_result)
            anthropic_pico = self.extract_pico_elements(anthropic_result)
            
            row_data = {
                'Study_ID': study_id,
                'Title': article.title or 'Unknown',
                
                'Population_OpenAI': openai_pico['Population'],
                'Population_Anthropic': anthropic_pico['Population'],
                'Intervention_OpenAI': openai_pico['Intervention'],
                'Intervention_Anthropic': anthropic_pico['Intervention'],
                'Comparison_OpenAI': openai_pico['Comparison'],
                'Comparison_Anthropic': anthropic_pico['Comparison'],
                'Outcomes_OpenAI': openai_pico['Outcomes'],
                'Outcomes_Anthropic': anthropic_pico['Outcomes'],
                'Time_OpenAI': openai_pico['Time'],
                'Time_Anthropic': anthropic_pico['Time'],
                'Study_Type_OpenAI': openai_pico['Study_Type'],
                'Study_Type_Anthropic': anthropic_pico['Study_Type'],
                
                'Inclusion_Criteria_OpenAI': 'Yes' if openai_result.get('criteria_evaluation', {}).get('meets_inclusion_criteria') else 'No',
                'Inclusion_Criteria_Anthropic': 'Yes' if anthropic_result.get('criteria_evaluation', {}).get('meets_inclusion_criteria') else 'No',
                'Inclusion_Reasoning_OpenAI': openai_result.get('criteria_evaluation', {}).get('inclusion_reasoning', 'Not provided'),
                'Inclusion_Reasoning_Anthropic': anthropic_result.get('criteria_evaluation', {}).get('inclusion_reasoning', 'Not provided'),
                
                'Exclusion_Criteria_OpenAI': 'Yes' if openai_result.get('criteria_evaluation', {}).get('violates_exclusion_criteria') else 'No',
                'Exclusion_Criteria_Anthropic': 'Yes' if anthropic_result.get('criteria_evaluation', {}).get('violates_exclusion_criteria') else 'No',
                'Exclusion_Reasoning_OpenAI': openai_result.get('criteria_evaluation', {}).get('exclusion_reasoning', 'Not provided'),
                'Exclusion_Reasoning_Anthropic': anthropic_result.get('criteria_evaluation', {}).get('exclusion_reasoning', 'Not provided'),
                
                'Final_Decision_OpenAI': openai_result.get('screening_decision', {}).get('final_decision', 'Unknown'),
                'Final_Decision_Anthropic': anthropic_result.get('screening_decision', {}).get('final_decision', 'Unknown'),
                'Confidence_OpenAI': f"{openai_result.get('screening_decision', {}).get('confidence_score', 0):.2f}",
                'Confidence_Anthropic': f"{anthropic_result.get('screening_decision', {}).get('confidence_score', 0):.2f}",
                'Reasoning_OpenAI': openai_result.get('screening_decision', {}).get('detailed_reasoning', 'Not provided'),
                'Reasoning_Anthropic': anthropic_result.get('screening_decision', {}).get('detailed_reasoning', 'Not provided'),
                
                'Agreement_Status': agreement_status,
                'Agreement_Score': f"{agreement_metrics['agreement_score']:.3f}",
                'Confidence_Delta': f"{agreement_metrics['confidence_delta']:.2f}",
                'Conflict_Detected': 'Yes' if agreement_metrics['conflict_detected'] else 'No',
                'Low_Confidence': 'Yes' if agreement_metrics['low_confidence'] else 'No',
                'Requires_Human_Review': 'Yes' if agreement_metrics['requires_human_review'] else 'No',
                'Final_Resolved_Decision': final_decision['decision'],
                'Final_Confidence': f"{final_decision['confidence']:.2f}",
                'Resolution_Method': final_decision['resolution_method'],
                'Primary_Provider': final_decision.get('primary_provider', 'N/A')
            }
            
            comparison_data.append(row_data)
        
        if not comparison_data:
            raise ValueError("No dual-LLM screening results found to export")
        
        df = pd.DataFrame(comparison_data)
        
        quality_metrics = self.calculate_quality_metrics(articles)
        
        import tempfile
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{project_name}_dual_llm_comparison_{timestamp}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        
        self._create_formatted_excel(df, filepath, quality_metrics)
        
        return filepath
    
    def _create_formatted_excel(self, df: pd.DataFrame, filepath: str, quality_metrics: Dict[str, Any]):
        """Create Excel file with color-coded formatting and quality metrics summary."""
        wb = Workbook()
        
        ws_main = wb.active
        ws_main.title = "Dual-LLM Comparison"
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_main.append(r)
        
        for cell in ws_main[1]:
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        for row_idx, row in enumerate(ws_main.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                column_name = ws_main.cell(row=1, column=col_idx).value
                
                if column_name and '_OpenAI' in column_name:
                    cell.fill = self.openai_fill
                elif column_name and '_Anthropic' in column_name:
                    cell.fill = self.anthropic_fill
                elif column_name == 'Agreement_Status':
                    if cell.value == 'Agreement':
                        cell.fill = self.agreement_fill
                    elif cell.value == 'Disagreement':
                        cell.fill = self.disagreement_fill
                elif column_name == 'Conflict_Detected':
                    if cell.value == 'Yes':
                        cell.fill = self.disagreement_fill
                elif column_name == 'Requires_Human_Review':
                    if cell.value == 'Yes':
                        cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        
        for column in ws_main.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_main.column_dimensions[column_letter].width = adjusted_width
        
        ws_summary = wb.create_sheet(title="Quality Metrics")
        
        summary_data = [
            ["Metric", "Value"],
            ["Total Evaluations", quality_metrics['total_evaluations']],
            ["", ""],
            ["Decision Distribution", ""],
            ["Include Decisions", quality_metrics['decision_distribution']['INCLUDE']],
            ["Exclude Decisions", quality_metrics['decision_distribution']['EXCLUDE']],
            ["Conflict Decisions", quality_metrics['decision_distribution']['CONFLICT']],
            ["Incomplete Decisions", quality_metrics['decision_distribution']['INCOMPLETE']],
            ["", ""],
            ["Quality Metrics", ""],
            ["Conflict Rate (%)", f"{quality_metrics['quality_metrics']['conflict_rate']:.2f}"],
            ["Human Review Rate (%)", f"{quality_metrics['quality_metrics']['human_review_rate']:.2f}"],
            ["Average Confidence", f"{quality_metrics['quality_metrics']['average_confidence']:.2f}"],
            ["Average Agreement (%)", f"{quality_metrics['quality_metrics']['average_agreement']:.2f}"]
        ]
        
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        for cell in ws_summary[1]:
            cell.font = self.header_font
            cell.alignment = self.center_alignment
        
        for row in ws_summary.iter_rows(min_row=2):
            if row[0].value in ["Decision Distribution", "Quality Metrics"]:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")  # Lavender
        
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max(max_length + 2, 15)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
